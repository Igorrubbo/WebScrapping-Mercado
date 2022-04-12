import dotenv
import pyodbc
import time
import os

from datetime import date
from openpyxl import load_workbook

from WSnagumo import webscrape_nagumo
from WSmercadolivre import webscrape_mercadolivre


# Checar tempo de execução do programa
start_time = time.time()

# Carregar variáveis de ambiente
dotenv.load_dotenv(dotenv.find_dotenv())

# Conexão com db do Microsoft Azure
AZURE_USER = os.getenv('AZURE_USER')
AZURE_PASSWD = os.getenv('AZURE_PASSWD')
AZURE_SERVER = os.getenv('AZURE_SERVER')

connection_data = ('Driver={ODBC Driver 18 for SQL Server};Server=%s;Database=Preços_DB;Uid=%s;Pwd=%s;Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30') % (AZURE_SERVER, AZURE_USER, AZURE_PASSWD)

connection_db = pyodbc.connect(connection_data)

mycursor = connection_db.cursor()   

# Data do dia de hoje em formato dd/mm/yy
today = date.today()
data_hoje = today.strftime("%Y-%m-%d")

# Carregar planilha e produtos com openpyxl
arquivo = r"lista de mercado.xlsx"
wb = load_workbook(arquivo)
ws = wb[wb.sheetnames[0]]
lista_produtos = []
coluna_link = ws['A'][1:]
for cell in coluna_link:
    lista_produtos.append(f'{cell.value}')

# Execução do webscrapping
webscrape_nagumo(lista_produtos, mycursor, data_hoje, ws)
webscrape_mercadolivre(lista_produtos, mycursor, data_hoje, ws)

# Fechar planilha excel e conexão com db
wb.save(filename = arquivo)
wb.close()
mycursor.close()
connection_db.close()

# Print tempo de execução do programa
print("--- %s seconds ---" % (time.time() - start_time))

