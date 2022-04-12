import os
import time
import dotenv
import pyodbc
from bs4 import BeautifulSoup
from datetime import date
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

"""
#Carregar variáveis de ambiente
dotenv.load_dotenv(dotenv.find_dotenv())

#Checar tempo de execução do programa
start_time = time.time()

#Data do dia de hoje em formato dd/mm/yy
today = date.today()
data_hoje = today.strftime("%Y-%m-%d")

#conexão com db do Microsoft Azure
AZURE_USER = os.getenv('AZURE_USER')
AZURE_PASSWD = os.getenv('AZURE_PASSWD')
AZURE_SERVER = os.getenv('AZURE_SERVER')

connection_data = ('Driver={ODBC Driver 18 for SQL Server};Server=%s;Database=teste;Uid=%s;Pwd=%s;Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30') % (AZURE_SERVER, AZURE_USER, AZURE_PASSWD)

connection_db = pyodbc.connect(connection_data)

#print('Connection Succesful')

mycursor = connection_db.cursor()

# Carregar planilha e produtos com openpyxl
arquivo = r"lista de mercado.xlsx"
wb = load_workbook(arquivo)
ws = wb[wb.sheetnames[0]]
lista_produtos = []
coluna_link = ws['A'][1:7]
for cell in coluna_link:
    lista_produtos.append(f'{cell.value}')
"""
# Instanciar um objeto Options e adicionar o argumento "--headless"
opts = webdriver.ChromeOptions()
opts.add_argument(" --headless")
opts.add_experimental_option('excludeSwitches', ['enable-logging'])

# Definir o local do webdriver e iniciar ele
chrome_driver = Service(r'C:\Program Files (x86)\Chromedriver\chromedriver.exe')
driver = webdriver.Chrome(options = opts, service = chrome_driver)

def webscrape_nagumo(lista_produtos, mycursor, data_hoje, ws_excel):
    for produto in lista_produtos:
        # Carregar a página HTML e realizar uma pausa para a página poder carregar
        url_dinamica_nagumo = r'http://www.nagumo.com.br/atibaia-lj32-atibaia-alvinopolis-avenida-prof-carlos-alberto-de-carvalho/busca/' + produto.replace(' ', '-')
        driver.get(url_dinamica_nagumo)
        time.sleep(2)

        # Começar a usar o Beautiful Soup para puxar os dados
        soup_file = driver.page_source
        site_nagumo = BeautifulSoup(soup_file, 'html.parser')
        itens = site_nagumo.findAll(class_ = 'product-grid-default product-grid-default-4 content-dailySale-list-products-product')

        # Função para procurar se o produto solicitado é o mesmo que o encontrado
        def checar_produto():
            nova_lista = produto.split()
            for palavra in nova_lista:
                resultado = (' ' + palavra.upper() + ' ') in (' ' + titulo_html.text.upper() + ' ')
                if resultado == False:
                    break
            return resultado

        if len(itens) < 7: #condição necessária para casos em que o site producrado tem menos de 7 produtos listados
            i = len(itens) - 1
        else:
            i = 7
        for item in itens[0:i]:
            titulo_auxiliar = item.find(class_ = 'txt-desc-product-itemtext-muted txt-desc-product-item')
            titulo_html = titulo_auxiliar.find(class_ = 'list-product-link') 
            titulo_nagumo = "'" + titulo_html.text + "'"
            link = item.find(class_ = 'list-product-link')
            link_nagumo = "'" + 'http://www.nagumo.com.br' + (link.text, link['href'])[1] + "'"
            # Verificação se existe oferta, caso exista usar o preço correto
            preco_total = item.find(class_ ='area-bloco-preco bloco-preco pr-0')
            preco_oferta = item.find(class_ ='preco-oferta')
            if preco_oferta:
                preco_nagumo = preco_oferta.text.split('R$')[1]
                preco_nagumo = "".join(preco_nagumo.split()).replace(",",".")
            else:
                preco_nagumo = preco_total.text.split('R$')[1]
                preco_nagumo = "".join(preco_nagumo.split()).replace(",",".")
            # Verificação se existe desconto, caso exista extrair a informação
            desconto_online = item.find(class_ = 'promotion-tip-text')
            outro_desconto = item.find(class_ = 'discount-tag')
            if desconto_online:
                desconto_nagumo = "'" + desconto_online.text + "'"
            elif outro_desconto:
                desconto_nagumo = "'" + outro_desconto.text + "'"
            else:
                desconto_nagumo = "'" + '0%' + "'"

            if checar_produto() != False:
                ws_excel['C' + str((lista_produtos.index(produto)+2))] = 'produto encontrado' # Descobrir quais produtos não são encontrados
                print('Título do produto: ' + titulo_nagumo)
                print('Link do produto: ' + link_nagumo)
                print('Preço do produto com desconto: R$' + preco_nagumo)
                print('Desconto aplicado: ' + desconto_nagumo)          
                print(data_hoje)    
                print('---------------------------------------------------')
                # Inserir os dados no banco de dados
                mycursor.execute("INSERT INTO preços_nagumo(produto, titulo, preço, desconto, dia, link) values(?, ?, ?, ?, ?, ?)", (produto, titulo_nagumo, preco_nagumo, desconto_nagumo, data_hoje, link_nagumo))
                mycursor.commit()
                break # Para o loop caso o produto tenha sido encontrado
            if item == itens[i - (i + 1)] and checar_produto == False: # Descobrir quais produtos não são encontrados
                ws_excel['B' + str((lista_produtos.index(produto)+2))] = url_dinamica_nagumo 
                print(f'Produto não encontrado o link é: {url_dinamica_nagumo}]')   
"""
webscrape_nagumo(lista_produtos, mycursor, data_hoje, ws)
wb.save(filename = arquivo)
wb.close()
mycursor.close()
connection_db.close()
print("--- %s seconds ---" % (time.time() - start_time))
"""