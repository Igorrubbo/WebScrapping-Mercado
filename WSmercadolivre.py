import requests
import os
import time
import dotenv
import pyodbc
from datetime import date
from bs4 import BeautifulSoup
from openpyxl import load_workbook
"""
#Carregar variáveis de ambiente
dotenv.load_dotenv(dotenv.find_dotenv())

#Checar tempo de execução do programa
start_time = time.time()

#Data do dia de hoje em formato yyyy/mm/dd
today = date.today()
data_hoje = today.strftime("%Y-%m-%d")


#conexão com db do Microsoft Azure
AZURE_USER = os.getenv('AZURE_USER')
AZURE_PASSWD = os.getenv('AZURE_PASSWD')
AZURE_SERVER = os.getenv('AZURE_SERVER')

connection_data = ('Driver={ODBC Driver 18 for SQL Server};Server=%s;Database=teste;Uid=%s;Pwd=%s;Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30') % (AZURE_SERVER, AZURE_USER, AZURE_PASSWD)

connection_db = pyodbc.connect(connection_data)

#print('Connection Succesful')

mycursor = connection_db.cursor()

# Carregar planilha e produtos com openpyxl
arquivo = r"lista de mercado.xlsx"
wb = load_workbook(arquivo)
ws = wb[wb.sheetnames[0]]
lista_produtos = []
coluna_link = ws['A'][1:4]
for cell in coluna_link:
    lista_produtos.append(f'{cell.value}')
"""

# Pegar informações do site
def webscrape_mercadolivre(lista_produtos, mycursor, data_hoje, ws_excel):
    for produto in lista_produtos:
        url_dinamica = r'https://lista.mercadolivre.com.br/supermercado/' + produto.replace(' ', '-') + '_Frete_Full_OrderId_PRICE_NoIndex_True#applied_filter_id%3Dshipping_highlighted_fulfillment%26applied_filter_name%3DTipo+de+envio%26applied_filter_order%3D2%26applied_value_id%3Dfulfillment%26applied_value_name%3DFull%26applied_value_order%3D1%26applied_value_results%3D33%26is_custom%3Dfalse'
        response = requests.get(url_dinamica)
        site = BeautifulSoup(response.text, 'html.parser')
        time.sleep(1) # Garantir que dê tempo para a página carregar

        # Método para pegar a classe do item (produtos diferentes tem classes diferentes)
        resultado = site.find('div', attrs={'class': 'ui-search-result__wrapper'})
        if not resultado: #caso não encontre nenhum resultado parar o loop
            break
        primeiro_resultado = list(resultado.children)[0]
        classe_dinamica = str(primeiro_resultado).split('>')[0][12:-1]

        # Função para procurar se o produto encontrado é o mesmo que o solicitado
        def checar_produto():
            nova_lista = produto.split()
            for palavra in nova_lista:
                resultado = (' ' + palavra.upper() + ' ') in (' ' + titulo_mercadolivre_html.text.upper() + ' ')
                if resultado == False:
                    break
            return resultado

        #Garantir que a classe encontrada não seja a classe errada
        i = 0
        while "advertisement" in classe_dinamica:
            i += 1
            aux_classe_dinamica = list(resultado[i].children)[0]
            classe_dinamica = str(aux_classe_dinamica).split('>')[0][12:-1]
        itens = site.findAll('div', attrs={'class': classe_dinamica})

        if len(itens) < 7: #condição necessária para casos em que o site producrado tem menos de 7 produtos listados
            i = len(itens) - 1
        else:
            i = 7    
        for item in itens[0:7]:
            # Dados dos itens encontrados
            titulo_mercadolivre_html = item.find('h2', attrs={'class': 'ui-search-item__title'})
            titulo_mercadolivre = "'" + titulo_mercadolivre_html.text + "'"
            link = item.find('a', attrs={'class': 'ui-search-link'})
            link_mercadolivre_html = link.text, link['href']
            link_mercadolivre = "'" + link_mercadolivre_html[1] + "'"
            preco = item.find('div', attrs={'class': "ui-search-price__second-line"})   #Se tiver desconto garantir que o preço com desconto seja pego
            preco_reais = preco.find('span', attrs={'class': "price-tag-fraction"})
            preco_centavos = preco.find('span', attrs={'class': "price-tag-cents"})
            # Condição para caso tenha centavos no preço
            if preco_centavos:
                preco_mercadolivre = preco_reais.text + '.' + preco_centavos.text
            else:
                preco_mercadolivre = preco_reais.text
            desconto_mercadolivre_html = item.find('span', attrs={'class': 'ui-search-price__discount'})  #Se tiver desconto recolher qual é o % de desconto
            if desconto_mercadolivre_html:
                desconto_mercadolivre = "'" + desconto_mercadolivre_html.text + "'"
            else:
                desconto_mercadolivre = "'0%'"

            # Printar os dados
            if checar_produto() != False:
                ws_excel['B' + str((lista_produtos.index(produto)+2))] = 'produto encontrado' # Descobrir quais produtos são encontrados
                print("Título do produto:", titulo_mercadolivre)
                print('Preço: R$ ' + preco_mercadolivre)
                print('Link do produto:' + link_mercadolivre)
                print(desconto_mercadolivre)
                print(data_hoje)
                print('---------------------------------------------------')
                # Inserir os dados no banco de dados
                mycursor.execute("INSERT INTO preços_mercadolivre(produto, titulo, preço, desconto, dia, link) values(?, ?, ?, ?, ?, ?)", (produto, titulo_mercadolivre, preco_mercadolivre, desconto_mercadolivre, data_hoje, link_mercadolivre))
                mycursor.commit()
                break    # Para o loop caso o produto tenha sido encontrado  
            if item == itens[i - (i + 1)] and checar_produto == False: # Descobrir quais produtos não são encontrados
                ws_excel['B' + str((lista_produtos.index(produto)+2))] = url_dinamica
                print(f'Produto não encontrado o link é: {url_dinamica}]')      
"""
webscrape_mercadolivre(lista_produtos, mycursor, data_hoje, ws)
wb.save(filename = arquivo)
wb.close()
mycursor.close()
connection_db.close()
print("--- %s seconds ---" % (time.time() - start_time))
"""