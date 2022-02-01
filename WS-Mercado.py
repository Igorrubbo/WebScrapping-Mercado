import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook


# Carregar planilha e produtos com openpyxl
arquivo = "Arquivos excel\lista de mercado.xlsx"
wb = load_workbook(arquivo)
ws = wb[wb.sheetnames[0]]
lista_produtos = []
coluna_link = ws['A'][2:]
for cell in coluna_link:
    lista_produtos.append(f'{cell.value}')
print(lista_produtos)

# Função para procurar se o produto solicitado é o mesmo que o encontrado
def checar_produto():
    nova_lista = produto.split()
    for palavra in nova_lista:
        resultado = (' ' + palavra.upper() + ' ') in (' ' + titulo.text.upper() + ' ')
        if resultado == False:
            return False


# Pegar informações do site
for produto in lista_produtos:
    url_dinamica = r'https://lista.mercadolivre.com.br/supermercado/' + produto.replace(" ", "-")
    response = requests.get(url_dinamica)
    site = BeautifulSoup(response.text, 'html.parser')

    # Método para pegar a classe do item (produtos diferentes tem classes diferentes)
    resultado = site.find('div', attrs={'class': 'ui-search-result__wrapper'})
    primeiro_resultado = list(resultado.children)[0]
    classe_dinamica = str(primeiro_resultado).split('>')[0][12:-1]
    ##Garantir que a classe encontrada não seja a classe de produtos de propaganda
    i = 0
    while "advertisement" in classe_dinamica:
        i += 1
        aux_classe_dinamica = list(resultado[i].children)[0]
        classe_dinamica = str(aux_classe_dinamica).split('>')[0][12:-1]
    itens = site.findAll('div', attrs={'class': classe_dinamica})

    for item in itens[0:3]:
        # Dados dos itens encontrados
        titulo = item.find('h2', attrs={'class': 'ui-search-item__title'})
        preco_reais = item.find('span', attrs={'class': "price-tag-fraction"})
        preco_centavos = item.find('span', attrs={'class': "price-tag-cents"})
        link = item.find('a', attrs={'class': 'ui-search-link'})
        link_corrigido = link.text, link['href']

        # Printar os dados
        if checar_produto() != False:
            print("Título do produto:", titulo.text)
            ws['C' + str((lista_produtos.index(produto)+3))] = titulo.text
            # Condição para caso não tenha centavos no preço
            if (preco_centavos):
                print('Preço: R$ ' + preco_reais.text + ',' + preco_centavos.text)
                ws['B' + str((lista_produtos.index(produto)+3))] = preco_reais.text + ',' + preco_centavos.text
            else:
                print('Preço: R$ ' + preco_reais.text)
                ws['B' + str((lista_produtos.index(produto)+3))] = preco_reais.text + ',' + preco_centavos.text
            print('Link do produto:' + link_corrigido[1])
            ws['D' + str((lista_produtos.index(produto)+3))] = link_corrigido[1]
            print()
        else:
            continue

wb.save(filename='Arquivos excel\lista de mercado.xlsx')
wb.close()
