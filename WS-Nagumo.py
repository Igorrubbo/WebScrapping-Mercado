from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service


import time
import requests

#Checar tempo de execução do programa
start_time = time.time()

# Carregar planilha e produtos com openpyxl
arquivo = r"lista de mercado.xlsx"
wb = load_workbook(arquivo)
ws = wb[wb.sheetnames[0]]
lista_produtos = []
coluna_link = ws['A'][2:]
for cell in coluna_link:
    lista_produtos.append(f'{cell.value}')
print(lista_produtos)

# Instanciar um objeto Options e adicionar o argumento "--headless"
opts = webdriver.ChromeOptions()
opts.add_argument(" --headless")
opts.add_experimental_option('excludeSwitches', ['enable-logging'])

# Definir o local do webdriver e iniciar ele
chrome_driver = Service(r'C:\Program Files (x86)\Chromedriver\chromedriver.exe')
driver = webdriver.Chrome(options = opts, service = chrome_driver)

def webscrape_nagumo():
    for produto in lista_produtos:
        # Carregar a página HTML e realizar uma pausa para a página poder carregar
        url_dinamica_nagumo = r'http://www.nagumo.com.br/atibaia-lj32-atibaia-alvinopolis-avenida-prof-carlos-alberto-de-carvalho/busca/' + produto.replace(' ', '-')
        driver.get(url_dinamica_nagumo)
        time.sleep(2.5)

        # Começar a usar o Beautiful Soup para puxar os dados
        soup_file = driver.page_source
        site_nagumo = BeautifulSoup(soup_file, 'html.parser')
        main = site_nagumo.findAll(class_ = 'product-grid-default product-grid-default-4 content-dailySale-list-products-product')

        # Função para procurar se o produto solicitado é o mesmo que o encontrado
        def checar_produto():
            nova_lista = produto.split()
            for palavra in nova_lista:
                resultado = (' ' + palavra.upper() + ' ') in (' ' + titulo_nagumo.text.upper() + ' ')
                if resultado == False:
                    return False

        for item in main[0:6]:

            titulo_auxiliar = item.find(class_ = 'txt-desc-product-itemtext-muted txt-desc-product-item')
            titulo_nagumo = titulo_auxiliar.find(class_ = 'list-product-link')
            link_nagumo = item.find(class_ = 'list-product-link')
            preco_total_nagumo = item.find(class_ ='area-bloco-preco bloco-preco pr-0')
            preco_oferta = item.find(class_ ='preco-oferta')
            desconto_online = item.find(class_ = 'promotion-tip-text')
            desconto_nagumo = item.find(class_ = 'discount-tag')

            if checar_produto() != False:
                print('Título do produto: ' + titulo_nagumo.text)
                ws['H' + str((lista_produtos.index(produto)+3))] = titulo_nagumo.text

                print('Link do produto: ' + 'http://www.nagumo.com.br' + (link_nagumo.text, link_nagumo['href'])[1])
                ws['I' + str((lista_produtos.index(produto)+3))] = 'http://www.nagumo.com.br' + (link_nagumo.text, link_nagumo['href'])[1]

                if preco_oferta:
                    print('Preço do produto com desconto: R$' + preco_total_nagumo.text.split('R$')[1])
                    ws['F' + str((lista_produtos.index(produto)+3))] = preco_oferta.text
                else:                
                    print('Preço do produto: ' + preco_total_nagumo.text)
                    ws['F' + str((lista_produtos.index(produto)+3))] = preco_total_nagumo.text
                if desconto_online:
                    print('Desconto online existente, código:' + desconto_online.text)
                    ws['G' + str((lista_produtos.index(produto)+3))] = desconto_online.text
                if desconto_nagumo:
                    print('Desconto aplicado: ' + desconto_nagumo.text)
                    ws['G' + str((lista_produtos.index(produto)+3))] = desconto_nagumo.text                    

                print('---------------------------------------------------')

                break
            else:
                ws['F' + str((lista_produtos.index(produto)+3))] = '---'
                ws['G' + str((lista_produtos.index(produto)+3))] = '---'
                ws['H' + str((lista_produtos.index(produto)+3))] = 'Produto não encontrado'
                ws['I' + str((lista_produtos.index(produto)+3))] = '---'

    wb.save(filename = arquivo)
    wb.close()

webscrape_nagumo()
print("--- %s seconds ---" % (time.time() - start_time))